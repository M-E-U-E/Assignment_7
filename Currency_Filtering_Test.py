import os
import logging
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(message)s')

# Initialize WebDriver
def init_driver():
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.implicitly_wait(10)
    return driver

# Ensure directory exists
def ensure_directory(path):
    if not os.path.exists(path):
        os.makedirs(path)

# Save DataFrame to Excel with auto-adjusted column widths and formatting
def save_with_auto_width(filepath, df):
    df.to_excel(filepath, index=False, engine='openpyxl')
    wb = load_workbook(filepath)
    ws = wb.active

    # Define styles for formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Adjust column widths and format headers
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception as e:
                logging.warning(f"Error calculating column width: {e}")
                pass
            cell.alignment = alignment
            cell.border = border
        ws.column_dimensions[col_letter].width = max_length + 5

    # Apply header formatting
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    wb.save(filepath)

# Test currency filter functionality
def test_currency_filter(driver, url, currency_list):
    logging.info(f"Starting Currency Filter Test for URL: {url}")
    results = []  # List to store individual test results for each currency

    try:
        driver.get(url)
        logging.info("Page loaded successfully.")

        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.TAG_NAME, "body")))

        # Scroll down to load all content
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)

        dropdown = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, ".select-wrap"))
        )
        driver.execute_script("arguments[0].click();", dropdown)

        options = dropdown.find_elements(By.CSS_SELECTOR, ".select-ul > li")
        logging.info(f"Found {len(options)} currency options.")

        if not options:
            logging.warning("No currency options found in the dropdown.")
            return []

        for currency_code, currency_data in currency_list.items():
            country = currency_data["Country"]
            symbol = currency_data["Symbol"]
            logging.info(f"Testing currency: {country} ({currency_code})")

            matching_option = None
            for option in options:
                if currency_code in option.get_attribute("data-currency-country"):
                    matching_option = option
                    break

            if not matching_option:
                logging.warning(f"Currency {country} ({currency_code}) not found in dropdown. Skipping.")
                continue

            try:
                driver.execute_script("arguments[0].click();", matching_option)
                WebDriverWait(driver, 10).until(
                    EC.text_to_be_present_in_element((By.ID, "js-currency-sort-footer"), symbol)
                )

                tiles = driver.find_elements(By.CLASS_NAME, "property-tiles")
                price_info = driver.find_elements(By.CLASS_NAME, "js-price-value")

                if not tiles or not price_info:
                    results.append({"Country Name": country, "Currency Name": currency_code, "Symbol": symbol, "Status": "Inconclusive", "Reason": "No property tiles or price info found to verify currency change"})
                    continue

                tile_texts = [tile.text for tile in tiles]
                price_texts = [price.text for price in price_info]

                logging.info(f"Currency: {symbol}")
                logging.info(f"Property Tiles Text: {tile_texts}")
                logging.info(f"Price Info Text: {price_texts}")

                expected_symbol = symbol.split()[0]
                symbol_match_tiles = any(expected_symbol in text for text in tile_texts)
                symbol_match_prices = any(expected_symbol in text for text in price_texts)

                if symbol_match_tiles and symbol_match_prices:
                    results.append({"Country Name": country, "Currency Name": currency_code, "Symbol": symbol, "Status": "Pass"})
                else:
                    results.append({"Country Name": country, "Currency Name": currency_code, "Symbol": symbol, "Status": "Fail"})

            except Exception as e:
                results.append({"Country Name": country, "Currency Name": currency_code, "Symbol": symbol, "Status": "Fail", "Reason": f"Error selecting currency {symbol}: {e}"})
                logging.error(f"Error for currency {symbol}: {str(e)}")

        return results

    except Exception as e:
        logging.error(f"Error during Currency Filter Test: {str(e)}")
        return []

# Main function
def main():
    url = "https://www.alojamiento.io/property/mall-of-i-stanbul-3/BC-6975002/"  # Replace with the actual URL
    output_dir = "test_results"
    ensure_directory(output_dir)

    output_results_xlsx = os.path.join(output_dir, "currency_test_results.xlsx")
    output_summary_xlsx = os.path.join(output_dir, "currency_test_summary.xlsx")

    currency_list = {
        "AE": {"Code": "AED", "Country": "AE", "Symbol": "\u062f.\u0625.", "Rate": 3.672985},
        "AU": {"Code": "AUD", "Country": "AU", "Symbol": "$", "Rate": 1.551424},
        "BD": {"Code": "BDT", "Country": "BD", "Symbol": "\u09f3", "Rate": 119.903132},
        "BE": {"Code": "EUR", "Country": "BE", "Symbol": "\u20ac", "Rate": 0.945147},
        "CA": {"Code": "CAD", "Country": "CA", "Symbol": "$", "Rate": 1.413566},
        "DE": {"Code": "EUR", "Country": "DE", "Symbol": "\u20ac", "Rate": 0.945147},
        "ES": {"Code": "EUR", "Country": "ES", "Symbol": "\u20ac", "Rate": 0.945147},
        "FR": {"Code": "EUR", "Country": "FR", "Symbol": "\u20ac", "Rate": 0.945147},
        "GB": {"Code": "GBP", "Country": "GB", "Symbol": "\u00a3", "Rate": 0.782878},
        "IE": {"Code": "GBP", "Country": "IE", "Symbol": "\u00a3", "Rate": 0.782878},
        "IT": {"Code": "EUR", "Country": "IT", "Symbol": "\u20ac", "Rate": 0.945147},
        "SG": {"Code": "SGD", "Country": "SG", "Symbol": "$", "Rate": 1.338978},
        "UK": {"Code": "GBP", "Country": "UK", "Symbol": "\u00a3", "Rate": 0.782878},
        "US": {"Code": "USD", "Country": "US", "Symbol": "$", "Rate": 1},
    }

    driver = init_driver()
    try:
        results = test_currency_filter(driver, url, currency_list)

        if results:
            df_results = pd.DataFrame(results)
            save_with_auto_width(output_results_xlsx, df_results)

        pass_count = len([res for res in results if res["Status"] == "Pass"])
        fail_count = len([res for res in results if res["Status"] == "Fail"])

        overall_status = "Pass" if fail_count == 0 else "Fail"
        comments = "All currencies passed successfully." if fail_count == 0 else f"{fail_count} currencies failed."

        summary_data = [{
            "Page URL": url,
            "Test Case": "Currency Filter Test",
            "Status": overall_status,
            "Comments": comments
        }]
        df_summary = pd.DataFrame(summary_data)

        save_with_auto_width(output_summary_xlsx, df_summary)
        logging.info(f"Test results saved to {output_results_xlsx}")
        logging.info(f"Test summary saved to {output_summary_xlsx}")

    except Exception as e:
        logging.error(f"An error occurred during execution: {e}")
    finally:
        driver.quit()

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        logging.info("Execution interrupted by user.")
