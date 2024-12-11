import os
import subprocess
import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(message)s')

# Ensure directory exists
def ensure_directory(path):
    if not os.path.exists(path):
        os.makedirs(path)

# Save DataFrame to Excel with auto-adjusted column widths and formatting
def save_with_auto_width(filepath, df, sheet_name):
    with pd.ExcelWriter(filepath, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        wb = writer.book
        ws = wb[sheet_name]

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Adjust column widths and format headers
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception as e:
                    logging.warning(f"Error calculating column width: {e}")
                    pass
                cell.alignment = alignment
                cell.border = border
            ws.column_dimensions[col_letter].width = max_length + 5

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment

# Run individual test scripts
def run_tests(test_scripts):
    for script in test_scripts:
        script_name = os.path.basename(script)
        logging.info(f"Running test: {script_name}")
        try:
            subprocess.run(["python", script], check=True)
        except subprocess.CalledProcessError as e:
            logging.error(f"Error running {script_name}: {e}")

# Consolidate all result files ending with "results.xlsx" into one file with separate sheets
def consolidate_results(result_dir, output_file):
    ensure_directory(result_dir)
    summary_data = []
    
    for file_name in os.listdir(result_dir):
        if file_name.endswith("results.xlsx"):
            file_path = os.path.join(result_dir, file_name)
            sheet_name = file_name.replace("_results.xlsx", "").title()

            # Load individual test results
            df = pd.read_excel(file_path)
            summary_data.append((df, sheet_name))

    # Save all sheets to the output file
    for idx, (df, sheet_name) in enumerate(summary_data):
        mode = 'w' if idx == 0 else 'a'  # Write mode for first sheet, append for others
        with pd.ExcelWriter(output_file, engine='openpyxl', mode=mode) as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            wb = writer.book
            ws = wb[sheet_name]

            # Adjust column widths and format headers
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception as e:
                        logging.warning(f"Error calculating column width: {e}")
                        pass
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.column_dimensions[col_letter].width = max_length + 5

            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="4F81BD")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    logging.info(f"Consolidated report saved to {output_file}")

# Main function
def main():
    test_scripts = [
        "Currency_Filtering_Test.py",
        "H1_Tag_Existence_Test.py",
        "HTML_Tag_Sequence_Test.py",
        "Image_Alt_Attribute_Test.py",
        "Scrape_Data_from_Script_Tag.py",
        "URL_Status_Code_Test.py",
    ]
    result_dir = "test_results"
    consolidated_report = os.path.join(result_dir, "report_model_details.xlsx")

    # Run all test scripts
    run_tests(test_scripts)

    # Consolidate results into one Excel file
    consolidate_results(result_dir, consolidated_report)

if __name__ == "__main__":
    main()
