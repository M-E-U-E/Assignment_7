import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# Ensure directory exists
def ensure_directory(path):
    if path and not os.path.exists(path):  # Check if the path is not empty
        os.makedirs(path)

# Save DataFrame to Excel with auto-adjusted column widths and formatting
def save_with_auto_width(filepath, df, sheet_name=None):
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

    # Format the sheet
    wb = load_workbook(filepath)
    ws = wb[sheet_name]

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Adjust column widths and format headers
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
            cell.alignment = alignment
            cell.border = border
        ws.column_dimensions[col_letter].width = max_length + 5  # Adjust column width
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    wb.save(filepath)

# Consolidate specified summary files into one report
def consolidate_summaries(file_list, report_file):
    # Remove the report file if it exists to prevent appending issues
    if os.path.exists(report_file):
        os.remove(report_file)

    summary_data = []

    # Combine all data into a single summary sheet
    for file_path in file_list:
        try:
            df = pd.read_excel(file_path)
            summary_data.append(df)
        except Exception as e:
            print(f"Error reading file {file_path}: {e}")

    if summary_data:
        combined_df = pd.concat(summary_data, ignore_index=True)

        # Fix alignment and add styles to the combined data
        save_with_auto_width(report_file, combined_df, sheet_name="Combined_Summary")
        print("Combined summary added as the first and only sheet.")
    else:
        print("No valid files found to consolidate.")

# Main function
def main():
    # List of specific summary files to consolidate
    file_list = [
        "test_results/currency_test_summary.xlsx",
        "test_results/h1_tag_summary.xlsx",
        "test_results/html_tag_summary.xlsx",
        "test_results/image_alt_summary.xlsx",
        "test_results/script_data_summary.xlsx",
        "test_results/url_status_summary.xlsx"
    ]

    # Output consolidated report file
    report_file = "test_results/report_model.xlsx"

    # Consolidate summaries
    consolidate_summaries(file_list, report_file)

if __name__ == "__main__":
    main()
